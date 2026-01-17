from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def apply_header_style(worksheet):
    """Apply styling to the header row."""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def apply_column_widths(worksheet):
    """Set optimal column widths."""
    column_widths = {
        'A': 20,  # date
        'B': 25,  # league
        'C': 20,  # team_1_name
        'D': 20,  # team_2_name
        'E': 25,  # bet_description
        'F': 18,  # bet_type
        'G': 12,  # bet_amount
        'H': 14,  # odds_team_1
        'I': 14,  # odds_team_2
        'J': 16,  # implied_prob_team_1
        'K': 16,  # implied_prob_team_2
        'L': 22,  # sportsbook
        'M': 12,  # status
        'N': 16,  # payout_amount
        'O': 20,  # notes
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width


def apply_borders(worksheet):
    """Apply borders to all cells with data."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                   min_col=1, max_col=15):
        for cell in row:
            cell.border = thin_border


def apply_data_formatting(worksheet):
    """Apply number formatting to data columns."""
    for row_num in range(2, worksheet.max_row + 1):
        # Format date/time (Column A)
        worksheet[f'A{row_num}'].alignment = Alignment(horizontal="center")
        
        # Format bet_amount as currency (Column G)
        worksheet[f'G{row_num}'].number_format = '$#,##0.00'
        worksheet[f'G{row_num}'].alignment = Alignment(horizontal="right")
        
        # Format odds (Columns H, I)
        worksheet[f'H{row_num}'].number_format = '0.00'
        worksheet[f'I{row_num}'].number_format = '0.00'
        worksheet[f'H{row_num}'].alignment = Alignment(horizontal="center")
        worksheet[f'I{row_num}'].alignment = Alignment(horizontal="center")
        
        # Format implied probabilities as percentage (Columns J, K)
        worksheet[f'J{row_num}'].number_format = '0.00"%"'
        worksheet[f'K{row_num}'].number_format = '0.00"%"'
        worksheet[f'J{row_num}'].alignment = Alignment(horizontal="center")
        worksheet[f'K{row_num}'].alignment = Alignment(horizontal="center")
        
        # Format payout_amount as currency (Column N)
        worksheet[f'N{row_num}'].number_format = '$#,##0.00'
        worksheet[f'N{row_num}'].alignment = Alignment(horizontal="right")
        
        # Center align status (Column M)
        worksheet[f'M{row_num}'].alignment = Alignment(horizontal="center")


def apply_status_colors(worksheet):
    """Apply color coding based on status (Column M)."""
    status_colors = {
        "Won": "C6EFCE",      # Light green
        "Lost": "FFC7CE",     # Light red
        "Pending": "FFFFCC",  # Light yellow
        "Voided": "E6E6E6",   # Light gray
        "Partial": "FFE699"   # Light orange
    }
    
    for row_num in range(2, worksheet.max_row + 1):
        status_cell = worksheet[f'M{row_num}']
        status_value = status_cell.value
        
        if status_value in status_colors:
            status_cell.fill = PatternFill(start_color=status_colors[status_value], 
                                          end_color=status_colors[status_value], 
                                          fill_type="solid")
            # Add font color for better contrast
            if status_value == "Lost":
                status_cell.font = Font(color="9C0006", bold=True)
            elif status_value == "Won":
                status_cell.font = Font(color="006100", bold=True)


def style_workbook(worksheet):
    """Apply all styling to the worksheet."""
    apply_header_style(worksheet)
    apply_column_widths(worksheet)
    apply_borders(worksheet)
    apply_data_formatting(worksheet)
    apply_status_colors(worksheet)
