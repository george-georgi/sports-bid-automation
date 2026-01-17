from datetime import datetime
from openpyxl import Workbook, load_workbook
from styles import style_workbook
from constants import FILE_NAME


def convert_odds_to_probability(odds):
    """Convert decimal odds to implied probability (0-100%)."""
    try:
        odds_float = float(odds)
        if odds_float <= 0:
            return 0
        probability = (1 / odds_float) * 100
        return round(probability, 2)
    except ValueError:
        return None


def initialize_file():
    """Create the XLSX with headers if it doesn't exist."""
    try:
        load_workbook(FILE_NAME)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bids"
        
        headers = [
            "Date",
            "League",
            "Team 1 Name",
            "Team 2 Name",
            "Bet Description",
            "Bet Type",
            "Bet Amount",
            "Odds Team 1",
            "Odds Team 2",
            "Implied Prob Team 1 %",
            "Implied Prob Team 2 %",
            "Sportsbook",
            "Status",
            "Payout Amount",
            "Notes"
        ]
        ws.append(headers)
        style_workbook(ws)
        wb.save(FILE_NAME)


def add_bid_to_excel(league, team_1_name, team_2_name, bet_description, bet_type, bet_amount, odds_team_1, odds_team_2, status, payout_amount, notes):
    """Append a new bid entry to the XLSX."""
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    prob_team_1 = convert_odds_to_probability(odds_team_1)
    prob_team_2 = convert_odds_to_probability(odds_team_2)
    sportsbook = "DraftKings Sportsbook"
    
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    
    ws.append([
        date,
        league,
        team_1_name,
        team_2_name,
        bet_description,
        bet_type,
        bet_amount,
        odds_team_1,
        odds_team_2,
        prob_team_1,
        prob_team_2,
        sportsbook,
        status,
        payout_amount,
        notes
    ])
    
    style_workbook(ws)
    wb.save(FILE_NAME)
