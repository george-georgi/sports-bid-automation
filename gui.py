import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
from styles import style_workbook

FILE_NAME = "weekly_bids.xlsx"

# Dropdown options
LEAGUES = ["NFL", "NBA", "MLB", "NHL", "Soccer", "College Football", "College Basketball", "MMA", "Boxing", "Other"]
BET_TYPES = ["Moneyline", "Spread", "Over/Under", "Parlay", "Prop", "Teaser", "Futures", "Tie No Bet"]
STATUSES = ["Won", "Lost"]
SOCCER_LEAGUES = ["Premier League (England)", "La Liga (Spain)", "Serie A (Italy)", "Bundesliga (Germany)", "Ligue 1 (France)", "Other"]

# Soccer teams by league
PREMIER_LEAGUE_TEAMS = ["Arsenal", "Aston Villa", "Bournemouth", "Brentford", "Brighton", "Chelsea", "Crystal Palace", "Everton", "Fulham", "Ipswich Town", "Leicester City", "Liverpool", "Manchester City", "Manchester United", "Newcastle United", "Nottingham Forest", "Southampton", "Tottenham", "West Ham", "Wolverhampton"]

LA_LIGA_TEAMS = ["Real Madrid", "Barcelona", "Atlético Madrid", "Real Sociedad", "Villarreal", "Athletic Bilbao", "Girona", "Real Betis", "Valencia", "Sevilla", "Las Palmas", "Getafe", "Rayo Vallecano", "Osasuna", "Real Valladolid", "Celta Vigo", "Leganés", "Alavés", "Mallorca", "Huesca"]

SERIE_A_TEAMS = ["Juventus", "Inter", "AC Milan", "Roma", "Lazio", "Napoli", "Atalanta", "Fiorentina", "Torino", "Sassuolo", "Sampdoria", "Monza", "Udinese", "Salernitana", "Frosinone", "Lecce", "Verona", "Como", "Cagliari", "Empoli"]

BUNDESLIGA_TEAMS = ["Bayern Munich", "Borussia Dortmund", "RB Leipzig", "Bayer Leverkusen", "Stuttgart", "Union Berlin", "Borussia Mönchengladbach", "Wolfsburg", "Eintracht Frankfurt", "Mainz", "Hoffenheim", "SC Freiburg", "Cologne", "Augsburg", "VfB Bochum", "Holstein Kiel", "St. Pauli", "Heidenheim"]

LIGUE_1_TEAMS = ["PSG", "Marseille", "Monaco", "Lyon", "Lille", "Lens", "Nice", "Rennes", "Nantes", "Toulouse", "Strasbourg", "Brest", "Saint-Étienne", "Montpellier", "Angers", "Le Havre", "Lorient", "Reims"]

SOCCER_TEAMS_MAP = {
    "Premier League (England)": PREMIER_LEAGUE_TEAMS,
    "La Liga (Spain)": LA_LIGA_TEAMS,
    "Serie A (Italy)": SERIE_A_TEAMS,
    "Bundesliga (Germany)": BUNDESLIGA_TEAMS,
    "Ligue 1 (France)": LIGUE_1_TEAMS,
}


def convert_odds_to_probability(odds):
    """Convert decimal odds to implied probability (0-100%)."""
    try:
        odds_float = float(odds)
        if odds_float <= 0:
            return 0
        probability = (1 / odds_float) * 100
        return round(probability, 2)
    except ValueError:
        return None


def initialize_file():
    """Create the XLSX with headers if it doesn't exist."""
    try:
        load_workbook(FILE_NAME)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bids"
        
        headers = [
            "Date",
            "League",
            "Team 1 Name",
            "Team 2 Name",
            "Bet Description",
            "Bet Type",
            "Bet Amount",
            "Odds Team 1",
            "Odds Team 2",
            "Implied Prob Team 1 %",
            "Implied Prob Team 2 %",
            "Sportsbook",
            "Status",
            "Payout Amount",
            "Notes"
        ]
        ws.append(headers)
        style_workbook(ws)
        wb.save(FILE_NAME)


def add_bid_to_excel(league, team_1_name, team_2_name, bet_description, bet_type, bet_amount, odds_team_1, odds_team_2, status, payout_amount, notes):
    """Append a new bid entry to the XLSX."""
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    prob_team_1 = convert_odds_to_probability(odds_team_1)
    prob_team_2 = convert_odds_to_probability(odds_team_2)
    sportsbook = "DraftKings Sportsbook"
    
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    
    ws.append([
        date,
        league,
        team_1_name,
        team_2_name,
        bet_description,
        bet_type,
        bet_amount,
        odds_team_1,
        odds_team_2,
        prob_team_1,
        prob_team_2,
        sportsbook,
        status,
        payout_amount,
        notes
    ])
    
    style_workbook(ws)
    wb.save(FILE_NAME)


class BidTrackerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Sports Bid Tracker")
        self.root.geometry("600x800")
        self.root.resizable(False, False)
        
        # Initialize Excel file
        initialize_file()
        
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Create main frame with scrollbar
        main_frame = ttk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title = ttk.Label(main_frame, text="Sports Bid Tracker", font=("Arial", 16, "bold"))
        title.pack(pady=10)
        
        # League Selection
        ttk.Label(main_frame, text="League *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.league_var = tk.StringVar()
        self.league_combo = ttk.Combobox(main_frame, textvariable=self.league_var, values=LEAGUES, state="readonly", width=50)
        self.league_combo.pack(fill=tk.X, pady=5)
        self.league_combo.bind("<<ComboboxSelected>>", self.on_league_change)
        
        # Soccer League Selection (hidden by default)
        self.soccer_frame = ttk.Frame(main_frame)
        ttk.Label(self.soccer_frame, text="Soccer League *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.soccer_league_var = tk.StringVar()
        self.soccer_league_combo = ttk.Combobox(self.soccer_frame, textvariable=self.soccer_league_var, values=SOCCER_LEAGUES, state="readonly", width=50)
        self.soccer_league_combo.pack(fill=tk.X, pady=5)
        self.soccer_league_combo.bind("<<ComboboxSelected>>", self.on_soccer_league_change)
        
        # Team 1 Selection
        self.team1_frame = ttk.Frame(main_frame)
        ttk.Label(self.team1_frame, text="Team 1 Name *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.team1_var = tk.StringVar()
        self.team1_combo = ttk.Combobox(self.team1_frame, textvariable=self.team1_var, state="readonly", width=50)
        self.team1_combo.pack(fill=tk.X, pady=5)
        
        # Team 2 Selection
        self.team2_frame = ttk.Frame(main_frame)
        ttk.Label(self.team2_frame, text="Team 2 Name *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.team2_var = tk.StringVar()
        self.team2_combo = ttk.Combobox(self.team2_frame, textvariable=self.team2_var, state="readonly", width=50)
        self.team2_combo.pack(fill=tk.X, pady=5)
        
        # Bet Description
        ttk.Label(main_frame, text="Bet Description *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.bet_desc_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.bet_desc_var, width=50).pack(fill=tk.X, pady=5)
        
        # Bet Type
        ttk.Label(main_frame, text="Bet Type *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.bet_type_var = tk.StringVar()
        ttk.Combobox(main_frame, textvariable=self.bet_type_var, values=BET_TYPES, state="readonly", width=50).pack(fill=tk.X, pady=5)
        
        # Bet Amount
        ttk.Label(main_frame, text="Bet Amount ($) *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.bet_amount_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.bet_amount_var, width=50).pack(fill=tk.X, pady=5)
        
        # Odds Team 1
        ttk.Label(main_frame, text="Odds Team 1 *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.odds1_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.odds1_var, width=50).pack(fill=tk.X, pady=5)
        
        # Odds Team 2
        ttk.Label(main_frame, text="Odds Team 2 *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.odds2_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.odds2_var, width=50).pack(fill=tk.X, pady=5)
        
        # Status
        ttk.Label(main_frame, text="Status *", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.status_var = tk.StringVar()
        ttk.Combobox(main_frame, textvariable=self.status_var, values=STATUSES, state="readonly", width=50).pack(fill=tk.X, pady=5)
        
        # Payout Amount
        ttk.Label(main_frame, text="Payout Amount ($)", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.payout_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.payout_var, width=50).pack(fill=tk.X, pady=5)
        
        # Notes
        ttk.Label(main_frame, text="Notes", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.notes_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.notes_var, width=50).pack(fill=tk.X, pady=5)
        
        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        ttk.Button(button_frame, text="Submit Bet", command=self.submit_bet).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        self.show_team_options()
    
    def on_league_change(self, event=None):
        """Handle league selection change."""
        self.show_team_options()
    
    def on_soccer_league_change(self, event=None):
        """Handle soccer league selection change."""
        soccer_league = self.soccer_league_var.get()
        if soccer_league in SOCCER_TEAMS_MAP:
            teams = SOCCER_TEAMS_MAP[soccer_league]
            self.team1_combo['values'] = teams
            self.team2_combo['values'] = teams
            self.team1_var.set('')
            self.team2_var.set('')
        else:
            self.team1_combo['values'] = []
            self.team2_combo['values'] = []
    
    def show_team_options(self):
        """Show appropriate team selection based on league."""
        league = self.league_var.get()
        
        # Hide both frames first
        self.soccer_frame.pack_forget()
        self.team1_frame.pack_forget()
        self.team2_frame.pack_forget()
        
        if league == "Soccer":
            self.soccer_frame.pack(fill=tk.X, pady=5, after=self.league_combo)
            self.team1_frame.pack(fill=tk.X, pady=5, after=self.soccer_frame)
            self.team2_frame.pack(fill=tk.X, pady=5, after=self.team1_frame)
            self.team1_combo['state'] = 'readonly'
            self.team2_combo['state'] = 'readonly'
        else:
            self.team1_frame.pack(fill=tk.X, pady=5, after=self.league_combo)
            self.team2_frame.pack(fill=tk.X, pady=5, after=self.team1_frame)
            self.team1_combo['values'] = []
            self.team2_combo['values'] = []
            self.team1_combo['state'] = 'normal'
            self.team2_combo['state'] = 'normal'
            self.team1_var.set('')
            self.team2_var.set('')
    
    def validate_inputs(self):
        """Validate all required fields."""
        errors = []
        
        if not self.league_var.get():
            errors.append("League is required")
        
        if self.league_var.get() == "Soccer" and not self.soccer_league_var.get():
            errors.append("Soccer League is required")
        
        if not self.team1_var.get():
            errors.append("Team 1 Name is required")
        
        if not self.team2_var.get():
            errors.append("Team 2 Name is required")
        
        if not self.bet_desc_var.get():
            errors.append("Bet Description is required")
        
        if not self.bet_type_var.get():
            errors.append("Bet Type is required")
        
        if not self.bet_amount_var.get():
            errors.append("Bet Amount is required")
        else:
            try:
                float(self.bet_amount_var.get())
            except ValueError:
                errors.append("Bet Amount must be a valid number")
        
        if not self.odds1_var.get():
            errors.append("Odds Team 1 is required")
        else:
            try:
                float(self.odds1_var.get())
            except ValueError:
                errors.append("Odds Team 1 must be a valid number")
        
        if not self.odds2_var.get():
            errors.append("Odds Team 2 is required")
        else:
            try:
                float(self.odds2_var.get())
            except ValueError:
                errors.append("Odds Team 2 must be a valid number")
        
        if not self.status_var.get():
            errors.append("Status is required")
        
        if self.payout_var.get():
            try:
                float(self.payout_var.get())
            except ValueError:
                errors.append("Payout Amount must be a valid number")
        
        return errors
    
    def submit_bet(self):
        """Submit the bet form."""
        errors = self.validate_inputs()
        
        if errors:
            messagebox.showerror("Validation Error", "\n".join(errors))
            return
        
        try:
            league = self.league_var.get()
            if league == "Soccer":
                league = self.soccer_league_var.get()
            
            add_bid_to_excel(
                league=league,
                team_1_name=self.team1_var.get(),
                team_2_name=self.team2_var.get(),
                bet_description=self.bet_desc_var.get(),
                bet_type=self.bet_type_var.get(),
                bet_amount=float(self.bet_amount_var.get()),
                odds_team_1=self.odds1_var.get(),
                odds_team_2=self.odds2_var.get(),
                status=self.status_var.get(),
                payout_amount=self.payout_var.get(),
                notes=self.notes_var.get()
            )
            
            messagebox.showinfo("Success", f"Bet added successfully!\n{self.team1_var.get()} vs {self.team2_var.get()}")
            self.clear_form()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add bet: {str(e)}")
    
    def clear_form(self):
        """Clear all form fields."""
        self.league_var.set('')
        self.soccer_league_var.set('')
        self.team1_var.set('')
        self.team2_var.set('')
        self.bet_desc_var.set('')
        self.bet_type_var.set('')
        self.bet_amount_var.set('')
        self.odds1_var.set('')
        self.odds2_var.set('')
        self.status_var.set('')
        self.payout_var.set('')
        self.notes_var.set('')
        self.show_team_options()


if __name__ == "__main__":
    root = tk.Tk()
    gui = BidTrackerGUI(root)
    root.mainloop()
